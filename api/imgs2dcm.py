import os
import random
import re
import gc
import copy
import string
from api.imgfile_info import imgfile_info
import numpy as np
from openpyxl import load_workbook

from PySide6.QtCore import Signal
from io import BytesIO
from uuid import uuid4
from PIL import ImageFile, Image
Image.MAX_IMAGE_PIXELS = None

# import pydicom
from pydicom import FileDataset
from pydicom.dataset import Dataset, FileMetaDataset
from pydicom.sequence import Sequence
from pydicom.uid import generate_uid, ExplicitVRLittleEndian, JPEGBaseline8Bit, JPEGLosslessSV1, JPEG2000Lossless, VLWholeSlideMicroscopyImageStorage, PYDICOM_IMPLEMENTATION_UID
from pydicom.encaps import encapsulate
from pydicom.encaps import encapsulate_extended

def parse_tag_file(tag_file):
    """
    解析tag文字檔案，獲取tag dict。
    
    Args:
        tag_file:  tag檔案的路徑
    Returns:
        tags: tag dict
    """
    tags = {}
    print(f"Loading metadata file:{tag_file}")
    # 檢查檔案附檔名是否為 ".xlsx"
    _, file_extension = os.path.splitext(tag_file)
    if file_extension == ".xlsx": # excel
        # 讀取Excel檔案
        wb = load_workbook(filename=tag_file)
        sheet = wb.active
        
        # 將除了第一行以外的所有內容讀取到tags底下
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, value = row
            tags[name] = value

    else: # 其他檔案當文字檔處理
        with open(tag_file, 'r', encoding='utf-8') as file:
            for line in file:
                line = line.strip()
                if line:
                    tag_line = line.split(': ', 1)
                    tag_name = tag_line[0]
                    tag_value = tag_line[1] if len(tag_line) > 1 else ""
                    tags[tag_name] = tag_value
    return tags

def generate_random_string(length=8):
    characters = string.ascii_uppercase + string.digits
    random_string = ''.join(random.choice(characters) for _ in range(length))
    return random_string

def dataset_from_tag_file(tag_file, level=-1):
    """
        解析tag文字檔並產生dataset。
        
        Args:
            tag_file: tag檔案的路徑
        Returns:
            ds: dataset
    """

    instance_uid = generate_uid()

    # 創建一個空的DICOM Dataset
    file_meta = FileMetaDataset()
    ds = Dataset()

    # 設置DICOM Dataset 的初始屬性以及file_meta
    file_meta.MediaStorageSOPClassUID = VLWholeSlideMicroscopyImageStorage # VL Whole Slide Microscopy Image Storage (WSI)
    file_meta.MediaStorageSOPInstanceUID = instance_uid
    file_meta.ImplementationClassUID = PYDICOM_IMPLEMENTATION_UID
    file_meta.FileMetaInformationVersion = b'\x00\x01'
    #file_meta.TransferSyntaxUID = ExplicitVRLittleEndian
    #file_meta.TransferSyntaxUID = JPEGBaseline8Bit  # JPEG Baseline
    file_meta.TransferSyntaxUID = JPEG2000Lossless  # JPEG Baseline
    file_meta.ImplementationVersionName = 'OFFIS_DCMTK_367'
    file_meta.FileMetaInformationGroupLength = len(file_meta)
    ds.file_meta = file_meta

    ds.is_little_endian = True
    ds.is_implicit_VR = False

    # 為了避免讀取不到資料，先將必要欄位填上預設隨機產生的值。
    ds.StudyInstanceUID = generate_uid() 
    ds.SeriesInstanceUID = generate_uid()
    ds.SOPClassUID = VLWholeSlideMicroscopyImageStorage # VL Whole Slide Microscopy Image Storage (WSI)
    ds.SOPInstanceUID =  file_meta.MediaStorageSOPInstanceUID
    ds.PatientID =  f'Patient{generate_random_string(8)}'
    ds.AccessionNumber = generate_random_string(8)
    ds.Modality = 'SM'
    ds.PatientName = 'Anonymous'
    ds.PatientSex = 'M'
    ds.PatientBirthDate = '19600101'
    ds.StudyDate = '20210623'
    ds.StudyTime = '181944.786966'

    ds.ImagedVolumeWidth = 0.00025
    ds.ImagedVolumeHeight = 0.00025
    ds.TotalPixelMatrixOriginSequence = Sequence([Dataset()])
    ds.TotalPixelMatrixOriginSequence[0].XOffsetInSlideCoordinateSystem = 0
    ds.TotalPixelMatrixOriginSequence[0].YOffsetInSlideCoordinateSystem = 0

    ds.ImageType = 'VOLUME'
    ds.LossyImageCompression = '01'
    ds.PlanarConfiguration = 0
    ds.SamplesPerPixel = 3
    ds.BitsAllocated = 8
    ds.BitsStored = 8
    ds.HighBit = 7
    ds.PhotometricInterpretation = 'RGB'
    ds.PixelRepresentation = 0
    
    # 解析標籤文字檔案，獲取標籤字典
    tags = {}
    if tag_file != "":
        tags = parse_tag_file(tag_file)

        # 將標籤填入 DICOM 資料集中
        ds.PatientID = tags.get('Patient ID',  ds.PatientID)
        ds.PatientName = tags.get('Patient Name', 'Anonmyous')
        ds.PatientBirthDate = tags.get('Patient Birth Date', '20000101')
        ds.PatientSex = tags.get('Patient Sex', 'O')
        ds.StudyInstanceUID = tags.get('Study Instance UID', ds.StudyInstanceUID)
        ds.AccessionNumber = tags.get('Accession Number', ds.AccessionNumber)
        ds.StudyDate = tags.get('Study Date', '')
        ds.StudyTime = tags.get('Study Time', '')
        ds.AttributeModificationDateTime = tags.get('Study Last Modified Date', '')
        #ds.StudyScheduledPathologistName = tags.get('Study Scheduled Pathologist Name', '')
        #ds.SecondPathologistName = tags.get('Second Pathologist Name', '')
        ds.SeriesInstanceUID = tags.get('Series Instance UID', ds.SeriesInstanceUID)
        ds.Manufacturer = tags.get('Manufacturer', '')
        ds.InstitutionName = tags.get('Institution Name', '')
        ds.InstitutionalDepartmentName = tags.get('Institutional Department Name', '')
        ds.DeviceSerialNumber = tags.get('Device Serial Number', '')
        ds.SoftwareVersions = tags.get('Software Versions', '')
        #ds.LastCalibrationDate = tags.get('Last Calibration Date', '')
        #ds.LastCalibrationTime = tags.get('Last Calibration Time', '')
        ds.SecondaryCaptureDeviceManufacturer = tags.get('Secondary Capture Device Manufacturer', '')
        ds.SecondaryCaptureDeviceManufacturerModelName = tags.get('Secondary Capture Device Manufacturer\'s Model Name', '')
        ds.SecondaryCaptureDeviceSoftwareVersions = tags.get('Secondary Capture Device Software Versions', '')
        ds.PixelSpacing = [float(x) for x in re.split(r'[,\\]',tags['Pixel Spacing'])]
        #ds.PixelSpacing = [float(x) for x in tags['Pixel Spacing'].split(', ')] re.split(r'[,\\]',tags['Pixel Spacing'])
        ds.LossyImageCompression = tags.get('Lossy Image Compression', '')
        ds.LossyImageCompressionRatio = tags.get('Lossy Image Compression Ratio', '')
        ds.LossyImageCompressionMethod = tags.get('Lossy Image Compression Method', '')
        #ds.Barcode = tags.get('Barcode', '')
        #ds.ScannerCalibrationStatus = tags.get('Scanner Calibration Status', '')
        ds.ContainerIdentifier = tags.get('Container Identifier', '')
        #ds.BlockIdentifier = tags.get('Block Identifier', '')
        #ds.PartIdentifier = tags.get('Part Identifier', '')
        #ds.StainName = tags.get('Stain Name', '')
        #ds.StainTypeName = tags.get('Stain Type Name', '')
        ds.InstitutionAddress = tags.get('Institution Address', '')
        ds.PhotometricInterpretation = tags.get('Photometric Interpretation', '')
        ds.SecondaryCaptureDeviceID = tags.get('Secondary Capture Device ID', '')
        ds.StudyID = tags.get('Study ID', '')

    return ds

def imgs2dcm(input_folder, file_info:imgfile_info, file_ext, update_signal:Signal, level=-1):
    """
        將一堆png依照命名規則的圖檔，加上文字檔的tag，生成multiframe DICOM WSI。
        命名規則: layer_{level}_region_{x_index}_{y_index}.png
        Args:
            input_folder: 輸入圖檔的文件夾
            output_file: 輸出的DICOM文件路徑
            tag_file: 標籤檔案的路徑
            file_ext: 圖檔的副檔名 (png or jpg)
    """
    # print(f"{file_info.output_folder[:1]}/{file_info.output_filename}|Generating Tags")
    # 從tag_file取得tag資料並產生dataset
    ds = dataset_from_tag_file(file_info.metadata_file)

    # print(f"{file_info.output_folder[:1]}/{file_info.output_filename}|Reading Image Files")
    # 提取所有圖像檔案的y和x坐標值
    y_values, x_values, img_files = [], [], []

    for file in os.listdir(input_folder):
        if file.endswith(file_ext):
            img_files.append(file)
            parts = file[:-4].split('_')
            y = parts[-2]
            x = parts[-1]
            y_values.append(int(y))
            x_values.append(int(x))
    pass

    # 計算最大的y和x值
    max_y = max(y_values)
    max_x = max(x_values)

    # 對圖像文件進行排序，確保按照正確的順序排列
    img_files.sort(key=lambda f: [int(d) for d in f[:-4].split('_')[-2:]])

    # 讀取第一張圖像獲取其大小
    first_image_path = os.path.join(input_folder, img_files[0])
    first_image = Image.open(first_image_path)
    target_size = first_image.size

    # 定義縮放比例
    scale_factor = 1  # 根據需要進行調整
    target_size = tuple([int(scale_factor*x) for x in target_size])

    # 計算調整後的TotalPixelMatrixRows和TotalPixelMatrixColumns
    total_rows = (max_y + 1) * target_size[1]
    total_columns = (max_x + 1) * target_size[0]

    # 創建一個空的Pixel Data列表
    pixel_data_list = []

    # 遍歷排好序的圖像文件
    for i, jpg_file in enumerate(img_files):
        # 顯示當前處理的圖像文件
        file_info.convert_status = f"將圖片資料儲存到dcm({i+1}/{len(img_files)})"
        #update_signal.emit(0)
        # print(f"{file_info.output_folder[:1]}/{file_info.output_filename}|Processing Image [{i+1}/{len(img_files)}]|[{jpg_file}]")

        # 讀取原始圖像
        loaded_image = Image.open(os.path.join(input_folder, jpg_file))
        loaded_image = loaded_image.convert("RGB")

        image_str_buf = BytesIO()

        #"""
        # # # 將圖像轉換為字節數據
        loaded_image.save(image_str_buf, format="JPEG2000", progressive=False) # 儲存為JPEG (JPEG2000 PIL 儲存有問題)
        pixel_data = image_str_buf.getvalue()

        # 添加到Pixel Data列表中
        pixel_data_list.append(pixel_data)
        #"""
        """
        # # # 將圖像轉換為字節數據
        loaded_image.save(image_str_buf, format="BMP", progressive=False) # 儲存為JPEG (JPEG2000 PIL 儲存有問題)
        pixel_data = image_str_buf.getvalue()

        # 添加到Pixel Data列表中
        pixel_data_list.append(pixel_data)
        """

        del loaded_image
        gc.collect()
    pass

    ### 設置DICOM數據集的相關屬性, Frames, Rows, Columns 數值 ###
    ds.NumberOfFrames = len(img_files)
    # print(f"{file_info.output_folder[:1]}/{file_info.output_filename}|[Number of frames]=[{ds.NumberOfFrames}]")

    #maxSlide = max(target_size[1], target_size[0])
    ##ds.Rows = maxSlide
    ##ds.Columns = maxSlide
    ds.Rows = target_size[1]
    ds.Columns = target_size[0]

    # 設置調整後的TotalPixelMatrixRows和TotalPixelMatrixColumns
    ds.TotalPixelMatrixRows = total_rows
    ds.TotalPixelMatrixColumns = total_columns
    
    # print(f"{file_info.output_folder[:1]}/{file_info.output_filename}|Processing Pixel Data")

    ### pydicom 填入到pixeldata (encapsulate the pixel data list) ###
    # ds.PixelData = encapsulate(pixel_data_list)
    out: tuple[bytes, bytes, bytes] = encapsulate_extended(pixel_data_list)
    ds.PixelData = out[0]
    ds.ExtendedOffsetTable = out[1]
    ds.ExtendedOffsetTableLengths = out[2]

    # 儲存DICOM檔案
    ds.save_as(f"{file_info.output_folder}/{file_info.output_filename}", write_like_original=False)
    print(f"{file_info.output_folder[:1]}/{file_info.output_filename}|Finished")
    print(file_info.output_folder)
